from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
try:
    from datetime import UTC
except ImportError:
    # For Python < 3.11
    from datetime import timezone
    UTC = timezone.utc
from dotenv import load_dotenv
from flask_dance.contrib.google import make_google_blueprint, google
import os
import pymysql
from functools import wraps
import uuid
import random
import string
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import base64
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import Flow
from PIL import Image
import io

# Load environment variables from .env file
load_dotenv()

# Create login blueprint
login_bp = Blueprint('login', __name__, template_folder='public')

# Admin credentials
ADMIN_USERNAME = 'Jonas'
ADMIN_PASSWORD = 'Cu@tureQues!2403'
ADMIN_EMAIL = 'kwajunhao@gmail.com'

# Database configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST'),
    'port': int(os.getenv('DB_PORT')),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'database': os.getenv('DB_NAME'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# Google OAuth setup (will be registered in main app only if configured)
google_bp = None
if os.getenv('GOOGLE_CLIENT_ID') and os.getenv('GOOGLE_CLIENT_SECRET'):
    google_bp = make_google_blueprint(
        client_id=os.getenv('GOOGLE_CLIENT_ID'),
        client_secret=os.getenv('GOOGLE_CLIENT_SECRET'),
        scope=[
            "openid",
            "https://www.googleapis.com/auth/userinfo.email", 
            "https://www.googleapis.com/auth/userinfo.profile"
        ],
        redirect_to='login.google_callback'
    )

# Database Functions
def get_db_connection():
    """Get database connection"""
    try:
        return pymysql.connect(**DB_CONFIG)
    except Exception as e:
        print(f"Database connection failed: {e}")
        return None

def init_database():
    """Initialize database tables"""
    connection = get_db_connection()
    if not connection:
        return False
    
    try:
        with connection.cursor() as cursor:
            # Create users table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id VARCHAR(36) PRIMARY KEY,
                    username VARCHAR(50) UNIQUE NOT NULL,
                    email VARCHAR(120) UNIQUE NOT NULL,
                    password VARCHAR(255),
                    profile_picture TEXT,
                    occupation VARCHAR(100),
                    birthday DATE,
                    labels TEXT,
                    online_start_time VARCHAR(10),
                    online_end_time VARCHAR(10),
                    is_google_user BOOLEAN DEFAULT FALSE,
                    email_verified BOOLEAN DEFAULT FALSE,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    last_login DATETIME,
                    INDEX idx_username (username),
                    INDEX idx_email (email)
                )
            """)
            
            # Create verification_codes table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS verification_codes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    email VARCHAR(120) NOT NULL,
                    code VARCHAR(6) NOT NULL,
                    expires DATETIME NOT NULL,
                    attempts INT DEFAULT 0,
                    user_data TEXT NOT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_email (email)
                )
            """)
            
            # Create reset_codes table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS reset_codes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    email VARCHAR(120) NOT NULL,
                    code VARCHAR(6) NOT NULL,
                    expires DATETIME NOT NULL,
                    attempts INT DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_email (email)
                )
            """)
            
        connection.commit()
        return True
    except Exception as e:
        print(f"Error creating database tables: {e}")
        return False
    finally:
        connection.close()

def find_user_by_username(username):
    """Find user by username"""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            return cursor.fetchone()
    except Exception as e:
        print(f"Error finding user by username: {e}")
        return None
    finally:
        connection.close()

def find_user_by_email(email):
    """Find user by email"""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
            return cursor.fetchone()
    except Exception as e:
        print(f"Error finding user by email: {e}")
        return None
    finally:
        connection.close()

def find_user_by_id(user_id):
    """Find user by ID"""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
            return cursor.fetchone()
    except Exception as e:
        print(f"Error finding user by ID: {e}")
        return None
    finally:
        connection.close()

def create_user(username, email, password, is_google_user=False, profile_picture=None, email_verified=False):
    """Create a new user"""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        with connection.cursor() as cursor:
            user_id = str(uuid.uuid4())
            hashed_password = generate_password_hash(password) if password else None
            
            cursor.execute("""
                INSERT INTO users (id, username, email, password, profile_picture, 
                                 is_google_user, email_verified, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (user_id, username, email, hashed_password, profile_picture,
                  is_google_user, email_verified, datetime.utcnow()))
            
            connection.commit()
            return {
                'id': user_id,
                'username': username,
                'email': email,
                'password': hashed_password,
                'profile_picture': profile_picture,
                'is_google_user': is_google_user,
                'email_verified': email_verified
            }
    except Exception as e:
        print(f"Error creating user: {e}")
        return None
    finally:
        connection.close()

def update_user(user_id, updates):
    """Update user data"""
    connection = get_db_connection()
    if not connection:
        return None
    
    try:
        with connection.cursor() as cursor:
            # Build dynamic UPDATE query
            set_clause = ', '.join([f"{key} = %s" for key in updates.keys()])
            values = list(updates.values()) + [user_id]
            
            cursor.execute(f"UPDATE users SET {set_clause} WHERE id = %s", values)
            connection.commit()
            
            # Return updated user
            cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
            return cursor.fetchone()
    except Exception as e:
        print(f"Error updating user: {e}")
        return None
    finally:
        connection.close()

def update_user_login_time(username):
    """Update user's last login time"""
    connection = get_db_connection()
    if not connection:
        return
    
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE users SET last_login = %s WHERE username = %s",
                (datetime.utcnow(), username)
            )
            connection.commit()
    except Exception as e:
        print(f"Error updating login time: {e}")
    finally:
        connection.close()

def delete_user_by_id(user_id):
    """Delete user by ID"""
    connection = get_db_connection()
    if not connection:
        return False
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
            connection.commit()
            return cursor.rowcount > 0
    except Exception as e:
        print(f"Error deleting user: {e}")
        return False
    finally:
        connection.close()

def load_users():
    """Load all users from database"""
    connection = get_db_connection()
    if not connection:
        return []
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
            return cursor.fetchall()
    except Exception as e:
        return []
    finally:
        connection.close()

def get_user_stats():
    """Get user statistics"""
    connection = get_db_connection()
    if not connection:
        return {
            'total_users': 0,
            'google_users': 0,
            'regular_users': 0,
            'users_today': 0,
            'users_this_week': 0
        }
    
    try:
        with connection.cursor() as cursor:
            # Total users
            cursor.execute("SELECT COUNT(*) as count FROM users")
            total_users = cursor.fetchone()['count']
            
            # Google users
            cursor.execute("SELECT COUNT(*) as count FROM users WHERE is_google_user = TRUE")
            google_users = cursor.fetchone()['count']
            
            regular_users = total_users - google_users
            
            # Users today
            cursor.execute("SELECT COUNT(*) as count FROM users WHERE DATE(created_at) = CURDATE()")
            users_today = cursor.fetchone()['count']
            
            # Users this week
            cursor.execute("SELECT COUNT(*) as count FROM users WHERE created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)")
            users_this_week = cursor.fetchone()['count']
            
            return {
                'total_users': total_users,
                'google_users': google_users,
                'regular_users': regular_users,
                'users_today': users_today,
                'users_this_week': users_this_week
            }
    except Exception as e:
        print(f"Error getting user stats: {e}")
        return {
            'total_users': 0,
            'google_users': 0,
            'regular_users': 0,
            'users_today': 0,
            'users_this_week': 0
        }
    finally:
        connection.close()

# Password Reset Functions

def generate_reset_code():
    """Generate a 6-digit reset code"""
    return ''.join(random.choices(string.digits, k=6))

def store_reset_code(email, code):
    """Store reset code with expiration time"""
    connection = get_db_connection()
    if not connection:
        print(f"DEBUG STORE: Database connection failed for {email}")
        return False
    
    try:
        with connection.cursor() as cursor:
            # Remove any existing reset code for this email
            cursor.execute("DELETE FROM reset_codes WHERE email = %s", (email,))
            deleted_count = cursor.rowcount
            print(f"DEBUG STORE: Deleted {deleted_count} existing reset codes for {email}")
            
            # Insert new reset code with proper timezone
            expires = datetime.now(UTC) + timedelta(minutes=15)
            print(f"DEBUG STORE: Inserting reset code - Email: {email}, Code: {code}, Expires: {expires}")
            cursor.execute("""
                INSERT INTO reset_codes (email, code, expires, attempts) 
                VALUES (%s, %s, %s, 0)
            """, (email, code, expires))
            
            connection.commit()
            
            # Verify the code was inserted
            cursor.execute("SELECT * FROM reset_codes WHERE email = %s", (email,))
            inserted_code = cursor.fetchone()
            if inserted_code:
                print(f"DEBUG STORE: Reset code committed successfully - ID: {inserted_code.get('id')}, Code: {inserted_code.get('code')}")
                return True
            else:
                print(f"DEBUG STORE: Failed to verify code insertion for {email}")
                return False
    except Exception as e:
        print(f"DEBUG STORE: Exception storing reset code for {email}: {e}")
        connection.rollback()
        return False
    finally:
        connection.close()

def verify_reset_code_helper(email, provided_code):
    """Verify reset code and check expiration"""
    connection = get_db_connection()
    if not connection:
        print(f"DEBUG VERIFY: Database connection failed for {email}")
        return False, "Database connection failed"
    
    try:
        with connection.cursor() as cursor:
            # Find reset code for this email
            cursor.execute("SELECT * FROM reset_codes WHERE email = %s ORDER BY created_at DESC LIMIT 1", (email,))
            code_data = cursor.fetchone()
            
            print(f"DEBUG VERIFY: Looking for reset code for email: {email}")
            print(f"DEBUG VERIFY: Code data found: {code_data is not None}")
            if code_data:
                print(f"DEBUG VERIFY: Stored code: '{code_data.get('code')}', Provided: '{provided_code}'")
                print(f"DEBUG VERIFY: Code expires at: {code_data.get('expires')}")
                print(f"DEBUG VERIFY: Current attempts: {code_data.get('attempts', 0)}")
            
            if not code_data:
                print(f"DEBUG VERIFY: No reset code found in database for {email}")
                return False, "No reset code found for this email"
            
            # Check expiration with proper timezone handling
            current_time = datetime.now(UTC)
            expires_time = code_data['expires']
            
            # Ensure expires_time has timezone info
            if hasattr(expires_time, 'tzinfo') and expires_time.tzinfo is None:
                expires_time = expires_time.replace(tzinfo=UTC)
            elif isinstance(expires_time, str):
                try:
                    expires_time = datetime.fromisoformat(expires_time.replace('Z', '+00:00'))
                except:
                    # If parsing fails, assume naive datetime is UTC
                    expires_time = datetime.fromisoformat(expires_time).replace(tzinfo=UTC)
            
            print(f"DEBUG VERIFY: Current time: {current_time}")
            print(f"DEBUG VERIFY: Expires time: {expires_time}")
            print(f"DEBUG VERIFY: Is expired: {current_time > expires_time}")
            
            if current_time > expires_time:
                print(f"DEBUG VERIFY: Code expired, deleting from database")
                cursor.execute("DELETE FROM reset_codes WHERE email = %s", (email,))
                connection.commit()
                return False, "Reset code has expired"
            
            # Check attempts limit
            attempts = code_data.get('attempts', 0)
            if attempts >= 3:
                print(f"DEBUG VERIFY: Too many attempts ({attempts}), code blocked")
                return False, "Too many failed attempts. Please request a new code."
            
            # Clean and compare codes
            stored_code = str(code_data['code']).strip()
            provided_code = str(provided_code).strip()
            
            print(f"DEBUG VERIFY: Comparing codes - stored: '{stored_code}', provided: '{provided_code}'")
            print(f"DEBUG VERIFY: Codes match: {stored_code == provided_code}")
            
            if stored_code == provided_code:
                print(f"DEBUG VERIFY: Code verified successfully, deleting from database")
                # Code matches - remove it and return success
                cursor.execute("DELETE FROM reset_codes WHERE email = %s", (email,))
                connection.commit()
                return True, "Code verified successfully"
            else:
                print(f"DEBUG VERIFY: Code mismatch, incrementing attempts")
                # Code doesn't match - increment attempts
                cursor.execute(
                    "UPDATE reset_codes SET attempts = COALESCE(attempts, 0) + 1 WHERE email = %s",
                    (email,)
                )
                connection.commit()
                return False, "Invalid reset code"
                
    except Exception as e:
        print(f"DEBUG VERIFY: Exception occurred: {str(e)}")
        return False, f"Error verifying code: {str(e)}"
    finally:
        connection.close()

def cleanup_expired_codes():
    """Remove expired reset codes"""
    connection = get_db_connection()
    if not connection:
        return
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM reset_codes WHERE expires < %s", (datetime.now(UTC),))
            cursor.execute("DELETE FROM verification_codes WHERE expires < %s", (datetime.now(UTC),))
            connection.commit()
    except Exception as e:
        print(f"Error cleaning up expired codes: {e}")
    finally:
        connection.close()

# Email Verification Functions

def generate_verification_code():
    """Generate a 6-digit verification code"""
    return ''.join(random.choices(string.digits, k=6))

def store_verification_code(email, code, user_data):
    """Store verification code with expiration time and user data"""
    connection = get_db_connection()
    if not connection:
        print(f"DEBUG STORE EMAIL: Database connection failed for {email}")
        return False
    
    try:
        with connection.cursor() as cursor:
            # Remove any existing verification code for this email
            cursor.execute("DELETE FROM verification_codes WHERE email = %s", (email,))
            deleted_count = cursor.rowcount
            print(f"DEBUG STORE EMAIL: Deleted {deleted_count} existing verification codes for {email}")
            
            # Insert new verification code with user data as JSON
            expires = datetime.now(UTC) + timedelta(minutes=15)
            import json
            print(f"DEBUG STORE EMAIL: Inserting verification code - Email: {email}, Code: {code}, Expires: {expires}")
            cursor.execute("""
                INSERT INTO verification_codes (email, code, expires, attempts, user_data) 
                VALUES (%s, %s, %s, 0, %s)
            """, (email, code, expires, json.dumps(user_data)))
            
            connection.commit()
            
            # Verify the code was inserted
            cursor.execute("SELECT * FROM verification_codes WHERE email = %s", (email,))
            inserted_code = cursor.fetchone()
            if inserted_code:
                print(f"DEBUG STORE EMAIL: Verification code committed successfully - ID: {inserted_code.get('id')}, Code: {inserted_code.get('code')}")
                return True
            else:
                print(f"DEBUG STORE EMAIL: Failed to verify code insertion for {email}")
                return False
    except Exception as e:
        print(f"DEBUG STORE EMAIL: Exception storing verification code for {email}: {e}")
        connection.rollback()
        return False
    finally:
        connection.close()

def verify_email_code_helper(email, provided_code):
    """Verify email verification code and check expiration"""
    connection = get_db_connection()
    if not connection:
        print(f"DEBUG EMAIL VERIFY: Database connection failed for {email}")
        return False, "Database connection failed", None
    
    try:
        with connection.cursor() as cursor:
            # Find verification code for this email
            cursor.execute("SELECT * FROM verification_codes WHERE email = %s ORDER BY created_at DESC LIMIT 1", (email,))
            code_data = cursor.fetchone()
            
            print(f"DEBUG EMAIL VERIFY: Looking for verification code for email: {email}")
            print(f"DEBUG EMAIL VERIFY: Code data found: {code_data is not None}")
            if code_data:
                print(f"DEBUG EMAIL VERIFY: Stored code: '{code_data.get('code')}', Provided: '{provided_code}'")
                print(f"DEBUG EMAIL VERIFY: Code expires at: {code_data.get('expires')}")
                print(f"DEBUG EMAIL VERIFY: Current attempts: {code_data.get('attempts', 0)}")
            
            if not code_data:
                print(f"DEBUG EMAIL VERIFY: No verification code found in database for {email}")
                return False, "No verification code found for this email", None
            
            # Check if code has expired (handle timezone properly)
            current_time = datetime.now(UTC)
            expires_time = code_data['expires']
            
            # If expires_time is naive (no timezone), assume it's UTC
            if hasattr(expires_time, 'tzinfo') and expires_time.tzinfo is None:
                expires_time = expires_time.replace(tzinfo=UTC)
            elif isinstance(expires_time, str):
                try:
                    expires_time = datetime.fromisoformat(expires_time.replace('Z', '+00:00'))
                except:
                    expires_time = datetime.fromisoformat(expires_time).replace(tzinfo=UTC)
            
            print(f"DEBUG EMAIL VERIFY: Current time: {current_time}")
            print(f"DEBUG EMAIL VERIFY: Expires time: {expires_time}")
            print(f"DEBUG EMAIL VERIFY: Is expired: {current_time > expires_time}")
            
            if current_time > expires_time:
                print(f"DEBUG EMAIL VERIFY: Code expired, deleting from database")
                # Remove expired code
                cursor.execute("DELETE FROM verification_codes WHERE email = %s", (email,))
                connection.commit()
                return False, "Verification code has expired", None
            
            # Check attempts limit
            attempts = code_data.get('attempts', 0)
            if attempts >= 3:
                print(f"DEBUG EMAIL VERIFY: Too many attempts ({attempts}), code blocked")
                return False, "Too many failed attempts. Please request a new code.", None
            
            # Check if code matches (ensure proper string comparison)
            stored_code = str(code_data['code']).strip()
            provided_code = str(provided_code).strip()
            
            print(f"DEBUG EMAIL VERIFY: Comparing codes - stored: '{stored_code}', provided: '{provided_code}'")
            print(f"DEBUG EMAIL VERIFY: Codes match: {stored_code == provided_code}")
            
            if stored_code != provided_code:
                print(f"DEBUG EMAIL VERIFY: Code mismatch, incrementing attempts")
                # Increment attempts
                cursor.execute(
                    "UPDATE verification_codes SET attempts = attempts + 1 WHERE email = %s",
                    (email,)
                )
                connection.commit()
                return False, "Invalid verification code", None
            
            # Code is valid - get user data and remove verification code
            import json
            user_data = json.loads(code_data['user_data']) if code_data['user_data'] else None
            cursor.execute("DELETE FROM verification_codes WHERE email = %s", (email,))
            connection.commit()
            print(f"DEBUG EMAIL VERIFY: Code verified successfully, returning user_data: {user_data}")
            return True, "Email verified successfully", user_data
    except Exception as e:
        print(f"DEBUG EMAIL VERIFY: Exception occurred: {str(e)}")
        return False, "Error verifying code", None
    finally:
        connection.close()

def cleanup_expired_verification_codes():
    """Remove expired verification codes"""
    cleanup_expired_codes()  # This function now handles both reset and verification codes

def send_verification_email(email, code):
    """Send email verification email using Gmail API or SMTP"""
    try:
        # Clean up expired codes first
        cleanup_expired_verification_codes()
        
        # Try Gmail API first, then fallback to SMTP
        success = send_verification_via_gmail_api(email, code)
        if not success:
            success = send_verification_via_smtp(email, code)
        
        return success
    except Exception as e:
        print(f"Error sending verification email: {e}")
        return False

def send_verification_via_gmail_api(email, code):
    """Send verification email using Gmail API"""
    try:
        # Gmail API credentials should be in environment or credentials file
        gmail_user = os.getenv('GMAIL_USER')
        
        if not gmail_user:
            return False
            
        # Create the email message
        msg = MIMEMultipart()
        msg['From'] = f"CultureQuest <{gmail_user}>"
        msg['To'] = email
        msg['Subject'] = "CultureQuest - Email Verification Code"
        
        # Email body with professional template
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #4a90e2;">CultureQuest</h1>
                    <p style="color: #666;">Email Verification</p>
                </div>
                
                <div style="background: #f9f9f9; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <h2 style="color: #4a90e2; margin-top: 0;">Verify Your Email Address</h2>
                    <p>Welcome to CultureQuest! Please use the verification code below to complete your account setup:</p>
                    
                    <div style="text-align: center; margin: 25px 0;">
                        <div style="font-size: 32px; font-weight: bold; color: #4a90e2; background: white; padding: 15px; border-radius: 8px; letter-spacing: 5px; display: inline-block; border: 2px dashed #4a90e2;">
                            {code}
                        </div>
                    </div>
                    
                    <p><strong>This code will expire in 15 minutes.</strong></p>
                    <p>If you didn't create an account with CultureQuest, please ignore this email.</p>
                </div>
                
                <div style="text-align: center; color: #666; font-size: 12px;">
                    <p>This email was sent by CultureQuest</p>
                    <p>Please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Note: Gmail API implementation would go here
        # For now, fall back to SMTP
        return False
        
    except Exception as e:
        print(f"Gmail API error: {e}")
        return False

def send_verification_via_smtp(email, code):
    """Send verification email using SMTP (Gmail)"""
    try:
        # Get Gmail credentials from environment variables
        gmail_user = os.getenv('GMAIL_USER')
        gmail_password = os.getenv('GMAIL_APP_PASSWORD')  # Use App Password, not regular password
        
        if not gmail_user or not gmail_password:
            print("Gmail credentials not found in environment variables")
            return False
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"CultureQuest <{gmail_user}>"
        msg['To'] = email
        msg['Subject'] = "CultureQuest - Email Verification Code"
        
        # Email body with professional template
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h1 style="color: #4a90e2;">CultureQuest</h1>
                    <p style="color: #666;">Email Verification</p>
                </div>
                
                <div style="background: #f9f9f9; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                    <h2 style="color: #4a90e2; margin-top: 0;">Verify Your Email Address</h2>
                    <p>Welcome to CultureQuest! Please use the verification code below to complete your account setup:</p>
                    
                    <div style="text-align: center; margin: 25px 0;">
                        <div style="font-size: 32px; font-weight: bold; color: #4a90e2; background: white; padding: 15px; border-radius: 8px; letter-spacing: 5px; display: inline-block; border: 2px dashed #4a90e2;">
                            {code}
                        </div>
                    </div>
                    
                    <p><strong>This code will expire in 15 minutes.</strong></p>
                    <p>If you didn't create an account with CultureQuest, please ignore this email.</p>
                </div>
                
                <div style="text-align: center; color: #666; font-size: 12px;">
                    <p>This email was sent by CultureQuest</p>
                    <p>Please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>"""
        
        msg.attach(MIMEText(body, 'html'))
        
        # Gmail SMTP settings
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, gmail_password)
        
        # Send email
        text = msg.as_string()
        server.sendmail(gmail_user, email, text)
        server.quit()
        
        print(f"Verification email sent successfully to {email}")
        return True
        
    except Exception as e:
        print(f"SMTP error: {e}")
        return False

def send_reset_email(email, code):
    """Send password reset email using Gmail API or SMTP"""
    try:
        # Clean up expired codes first
        cleanup_expired_codes()
        
        # Try Gmail API first, then fallback to SMTP
        success = send_via_gmail_api(email, code)
        if not success:
            success = send_via_smtp(email, code)
        
        return success
    except Exception as e:
        print(f"Error sending reset email: {e}")
        return False

def send_via_gmail_api(email, code):
    """Send email using Gmail API"""
    try:
        # Gmail API credentials should be in environment or credentials file
        gmail_user = os.getenv('GMAIL_USER')
        
        if not gmail_user:
            return False
            
        # Create the email message
        msg = MIMEMultipart()
        msg['From'] = f"CultureQuest <{gmail_user}>"
        msg['To'] = email
        msg['Subject'] = "CultureQuest - Password Reset Code"
        
        # Email body
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h2 style="color: #4a90e2;">🔑 CultureQuest Password Reset</h2>
                </div>
                
                <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                    <h3 style="margin-top: 0; color: #2c3e50;">Your verification code is:</h3>
                    <div style="font-size: 32px; font-weight: bold; color: #4a90e2; text-align: center; background: white; padding: 15px; border-radius: 8px; letter-spacing: 8px; margin: 20px 0;">
                        {code}
                    </div>
                    <p style="color: #666; font-size: 14px; text-align: center;">
                        This code will expire in 15 minutes for security reasons.
                    </p>
                </div>
                
                <div style="background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <p style="margin: 0; color: #856404;">🛡️ <strong>Security Notice:</strong></p>
                    <ul style="color: #856404; margin: 10px 0;">
                        <li>Never share this code with anyone</li>
                        <li>CultureQuest will never ask for this code via phone or email</li>
                        <li>If you didn't request this reset, you can safely ignore this email</li>
                    </ul>
                </div>
                
                <div style="text-align: center; padding-top: 20px; border-top: 1px solid #eee; color: #666; font-size: 12px;">
                    <p>This email was sent from CultureQuest Password Reset System</p>
                    <p>© 2024 CultureQuest. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Try to use Gmail API (requires proper setup)
        # For now, fallback to SMTP
        return False
        
    except Exception as e:
        print(f"Gmail API error: {e}")
        return False

def send_via_smtp(email, code):
    """Send email using SMTP (Gmail)"""
    try:
        # Get Gmail credentials from environment variables
        gmail_user = os.getenv('GMAIL_USER')
        gmail_password = os.getenv('GMAIL_APP_PASSWORD')  # Use App Password, not regular password
        
        if not gmail_user or not gmail_password:
            print("Gmail credentials not found in environment variables")
            return False
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"CultureQuest <{gmail_user}>"
        msg['To'] = email
        msg['Subject'] = "CultureQuest - Password Reset Code"
        
        # Email body
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <h2 style="color: #4a90e2;">🔑 CultureQuest Password Reset</h2>
                </div>
                
                <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                    <h3 style="margin-top: 0; color: #2c3e50;">Your verification code is:</h3>
                    <div style="font-size: 32px; font-weight: bold; color: #4a90e2; text-align: center; background: white; padding: 15px; border-radius: 8px; letter-spacing: 8px; margin: 20px 0;">
                        {code}
                    </div>
                    <p style="color: #666; font-size: 14px; text-align: center;">
                        This code will expire in 15 minutes for security reasons.
                    </p>
                </div>
                
                <div style="background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <p style="margin: 0; color: #856404;">🛡️ <strong>Security Notice:</strong></p>
                    <ul style="color: #856404; margin: 10px 0;">
                        <li>Never share this code with anyone</li>
                        <li>CultureQuest will never ask for this code via phone or email</li>
                        <li>If you didn't request this reset, you can safely ignore this email</li>
                    </ul>
                </div>
                
                <div style="text-align: center; padding-top: 20px; border-top: 1px solid #eee; color: #666; font-size: 12px;">
                    <p>This email was sent from CultureQuest Password Reset System</p>
                    <p>© 2024 CultureQuest. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Gmail SMTP setup
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, gmail_password)
        
        # Send email
        text = msg.as_string()
        server.sendmail(gmail_user, email, text)
        server.quit()
        
        print(f"Password reset email sent successfully to {email}")
        return True
        
    except Exception as e:
        print(f"SMTP error: {e}")
        return False

# Admin decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'is_admin' not in session or not session['is_admin']:
            return redirect(url_for('login.login_page'))
        
        # Check if 2FA is pending (admin logged in but 2FA not completed)
        # Only redirect if we're not already on the 2FA page to prevent infinite loops
        if session.get('admin_2fa_pending'):
            current_endpoint = request.endpoint
            if current_endpoint != 'login.admin_2fa_page' and current_endpoint != 'login.admin_2fa_verify':
                return redirect(url_for('login.admin_2fa_page'))
        
        # Execute the protected function
        response = f(*args, **kwargs)
        
        # Add cache control headers to prevent browser caching
        if hasattr(response, 'headers'):
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        
        return response
    return decorated_function

# Routes
@login_bp.route('/')
def landing_page():
    """Landing page"""
    return render_template('landing.html')

@login_bp.route('/login')
def login_page():
    """Login page"""
    return render_template('login.html')

# Only register this fallback route if Google OAuth is not configured
if not os.getenv('GOOGLE_CLIENT_ID') or not os.getenv('GOOGLE_CLIENT_SECRET'):
    @login_bp.route('/auth/google')
    def google_login_not_configured():
        """Handle Google login when not configured"""
        return redirect(url_for('login.login_page'))



@login_bp.route('/signup', methods=['GET'])
def signup_page():
    """Regular signup page"""
    return render_template('signup.html')

@login_bp.route('/signup', methods=['POST'])
def signup():
    """Handle regular signup form submission"""
    # Check if request is JSON or form data
    if request.is_json:
        data = request.get_json()
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
    else:
        # Handle form data (this is what your HTML form sends)
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
    
    # Validate required fields
    if not username or not email or not password:
        if request.is_json:
            return jsonify({'success': False, 'message': 'All fields are required'})
        else:
            return redirect(url_for('login.signup_page'))
    
    # Restrict admin username
    if username.lower() == 'admin':
        if request.is_json:
            return jsonify({'success': False, 'message': 'Username "admin" is not allowed'})
        else:
            return redirect(url_for('login.signup_page'))
    
    # Check if user already exists
    if find_user_by_email(email):
        if request.is_json:
            return jsonify({'success': False, 'message': 'Email already registered', 'show_email_error': True})
        else:
            flash('Email already registered', 'error')
            return redirect(url_for('login.signup_page'))
    
    if find_user_by_username(username):
        if request.is_json:
            return jsonify({'success': False, 'message': 'Username already taken'})
        else:
            return redirect(url_for('login.signup_page'))
    
    # Generate verification code and store user data temporarily
    verification_code = generate_verification_code()
    user_data = {
        'username': username,
        'email': email,
        'password': password,
        'is_google_user': False,
        'profile_picture': None
    }
    
    # Store verification code with user data
    print(f"DEBUG: Storing verification code for {email}")
    if store_verification_code(email, verification_code, user_data):
        print(f"DEBUG: Verification code stored successfully")
        # Send verification email
        if send_verification_email(email, verification_code):
            print(f"DEBUG: Verification email sent successfully, redirecting to verify page")
            # Check if this is an AJAX request (from signup.html JavaScript)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'success': True, 
                    'message': 'Verification email sent',
                    'redirect_url': url_for('login.verify_email_page', email=email)
                })
            elif request.is_json:
                return jsonify({'success': True, 'message': 'Verification email sent'})
            else:
                return redirect(url_for('login.verify_email_page', email=email))
        else:
            if request.is_json:
                return jsonify({'success': False, 'message': 'Failed to send verification email'})
            else:
                return redirect(url_for('login.signup_page'))
    else:
        if request.is_json:
            return jsonify({'success': False, 'message': 'Error processing signup'})
        else:
            return redirect(url_for('login.signup_page'))

@login_bp.route('/signup2', methods=['GET'])
def signup2_page():
    """Google user signup page (after Google OAuth)"""
    if 'google_email' not in session:
        return redirect(url_for('login.login_page'))
    
    # Clear any previous flash messages to avoid showing irrelevant messages
    session.pop('_flashes', None)
    
    email = session.get('google_email')
    return render_template('signup2.html', email=email)

@login_bp.route('/signup2', methods=['POST'])
def signup2():
    """Handle Google user signup form submission"""
    # Check if it's an AJAX request (from JavaScript)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if 'google_email' not in session:
        if is_ajax:
            return jsonify({'error': 'Session expired. Please log in again.'}), 401
        return redirect(url_for('login.login_page'))
    
    # Get form data
    username = request.form.get('username')
    password = request.form.get('password')
    
    email = session.get('google_email')
    # Use default profile picture instead of Google profile picture
    profile_picture = 'default_profile.png'
    
    if not username or not password:
        if is_ajax:
            return jsonify({'error': 'Username and password are required'}), 400
        return redirect(url_for('login.signup2_page'))
    
    # Restrict admin username
    if username.lower() == 'admin':
        if is_ajax:
            return jsonify({'error': 'Username "admin" is not allowed'}), 400
        return redirect(url_for('login.signup2_page'))
    
    # Check if username already exists
    if find_user_by_username(username):
        if is_ajax:
            return jsonify({'error': 'Username already taken'}), 400
        return redirect(url_for('login.signup2_page'))
    
    # Check if email already exists
    if find_user_by_email(email):
        if is_ajax:
            return jsonify({'error': 'Email already registered'}), 400
        return redirect(url_for('login.signup2_page'))
    
    # Create new Google user
    new_user = create_user(username, email, password, is_google_user=True, profile_picture=profile_picture, email_verified=True)
    
    if new_user:
        # Set session for login
        session['username'] = username
        session['email'] = email
        session['profile_picture'] = 'default_profile.png'
        session['user_id'] = new_user['id']
        session['is_admin'] = False
        
        # Clear temporary Google session data
        session.pop('google_email', None)
        session.pop('google_name', None)
        session.pop('google_profile_picture', None)
        
        if is_ajax:
            return jsonify({
                'success': True,
                'message': 'Account created successfully!',
                'redirect_url': url_for('login.landing_page')
            })
        return redirect(url_for('login.landing_page'))
    else:
        if is_ajax:
            return jsonify({'error': 'Error creating account. Please try again.'}), 500
        return redirect(url_for('login.signup2_page'))

@login_bp.route('/login', methods=['POST'])
def login():
    """Handle login form submission"""
    username = request.form.get('username')
    password = request.form.get('password')

    if not username or not password:
        flash('Error. Try again.', 'error')
        return redirect(url_for('login.login_page'))

    # Check for admin login
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session['username'] = username  
        session['is_admin'] = True
        session['admin_2fa_pending'] = True  # Set flag for 2FA requirement
        return redirect(url_for('login.admin_2fa_page'))

    # Check regular user login (including Google users with passwords)
    user = find_user_by_username(username)

    if user and user.get('password') and check_password_hash(user['password'], password):
        # Update last login time
        update_user_login_time(username)
        
        session['username'] = username
        session['email'] = user['email']
        session['profile_picture'] = user.get('profile_picture')
        session['user_id'] = user['id']
        session['is_admin'] = False
        
        return redirect(url_for('login.landing_page'))
    
    # Check if it's a Google user trying to login manually by email
    user_by_email = find_user_by_email(username)
    if user_by_email and user_by_email.get('is_google_user') and user_by_email.get('password') and check_password_hash(user_by_email['password'], password):
        # Update last login time
        update_user_login_time(user_by_email['username'])
        
        session['username'] = user_by_email['username']
        session['email'] = user_by_email['email']
        session['profile_picture'] = user_by_email.get('profile_picture')
        session['user_id'] = user_by_email['id']
        session['is_admin'] = False
        
        return redirect(url_for('login.landing_page'))
    
    # Invalid credentials - show error message
    flash('Error. Try again.', 'error')
    return redirect(url_for('login.login_page'))

@login_bp.route('/profile')
def profile():
    """User profile page"""
    if 'username' in session:
        # Check if user is admin
        if session.get('is_admin', False):
            # Redirect admin to admin dashboard
            return redirect(url_for('login.admin_dashboard'))
        
        # Handle regular users
        username = session['username']
        user = find_user_by_username(username)
        if user:
            return render_template('profile.html', 
                                 username=username, 
                                 email=user['email'],
                                 profile_picture=user.get('profile_picture'), 
                                 occupation=user.get('occupation'), 
                                 birthday=user.get('birthday'), 
                                 labels=user.get('labels'),
                                 online_start_time=user.get('online_start_time'), 
                                 online_end_time=user.get('online_end_time'))
    return redirect(url_for('login.login_page'))

@login_bp.route('/update-profile', methods=['POST'])
def update_profile():
    """Update user profile"""
    if 'username' not in session or session.get('is_admin', False):
        return redirect(url_for('login.login_page'))

    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('login.login_page'))

    # Get form data
    online_start_time = request.form.get('online_start_time')
    online_end_time = request.form.get('online_end_time')
    occupation = request.form.get('occupation')
    birthday = request.form.get('birthday')
    labels = request.form.get('labels')

    updates = {}
    
    # Format time fields
    if online_start_time and online_end_time:
        try:
            start_time_obj = datetime.strptime(online_start_time, '%H:%M')
            end_time_obj = datetime.strptime(online_end_time, '%H:%M')
            formatted_start_time = start_time_obj.strftime('%I:%M%p').lower()
            formatted_end_time = end_time_obj.strftime('%I:%M%p').lower()
            updates['online_start_time'] = formatted_start_time
            updates['online_end_time'] = formatted_end_time
        except ValueError:
            return redirect(url_for('login.profile'))
    else:
        updates['online_start_time'] = None
        updates['online_end_time'] = None

    # Update other fields
    if occupation:
        updates['occupation'] = occupation.strip()
    if birthday:
        updates['birthday'] = birthday
    if labels:
        updates['labels'] = labels.strip()

    # Save updates
    update_user(user_id, updates)
    
    return redirect(url_for('login.profile'))

@login_bp.route('/upload-profile-picture', methods=['POST'])
def upload_profile_picture():
    """Handle profile picture upload"""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'}), 401
    
    if 'profile_picture' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    
    file = request.files['profile_picture']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    # Check file type and size
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    max_size = 5 * 1024 * 1024  # 5MB
    
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Invalid file type. Please use PNG, JPG, JPEG, or GIF'}), 400
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > max_size:
        return jsonify({'success': False, 'message': 'File size too large. Maximum 5MB allowed'}), 400
    
    try:
        # Create uploads directory if it doesn't exist
        upload_dir = os.path.join('static', 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        unique_filename = f"{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        filepath = os.path.join(upload_dir, unique_filename)
        
        # Process and resize image
        image = Image.open(file)
        
        # Convert RGBA to RGB if necessary
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        
        # Resize image to reasonable size (max 400x400)
        image.thumbnail((400, 400), Image.Resampling.LANCZOS)
        
        # Save processed image
        image.save(filepath, 'JPEG', quality=85, optimize=True)
        
        # Update database with new profile picture
        user_id = session.get('user_id')
        relative_path = f"uploads/{unique_filename}"
        
        if update_user(user_id, {'profile_picture': relative_path}):
            # Update session
            session['profile_picture'] = relative_path
            
            # Clean up old profile picture if it exists and isn't default
            old_pic = session.get('old_profile_picture')
            if old_pic and old_pic != 'default_profile.png' and old_pic.startswith('uploads/'):
                old_filepath = os.path.join('static', old_pic)
                if os.path.exists(old_filepath):
                    try:
                        os.remove(old_filepath)
                    except:
                        pass  # Ignore cleanup errors
            
            return jsonify({
                'success': True, 
                'message': 'Profile picture updated successfully',
                'image_url': f'/static/{relative_path}'
            })
        else:
            # Remove uploaded file if database update failed
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'message': 'Failed to update database'}), 500
            
    except Exception as e:
        print(f"Error uploading profile picture: {e}")
        return jsonify({'success': False, 'message': 'Error processing image'}), 500

@login_bp.route('/logout')
def logout():
    """Logout user"""
    # Check if this was an admin logout
    was_admin = session.get('is_admin', False)
    admin_username = session.get('username') if was_admin else None
    
    # Clear all session data
    session.clear()
    
    # If it was an admin, set a flag to require 2FA on any back button usage
    if was_admin and admin_username == ADMIN_USERNAME:
        session['admin_logged_out'] = True
        session['admin_username'] = admin_username
        print(f"DEBUG LOGOUT: Admin {admin_username} logged out, 2FA will be required on back button")
    
    return redirect(url_for('login.login_page'))

@login_bp.route('/auth/google/callback')
def google_callback():
    """Handle Google OAuth callback"""
    if not google.authorized:
        return redirect(url_for('login.login_page'))
    
    try:
        # Get user data from Google
        user_info = google.get('/oauth2/v2/userinfo')
        
        if not user_info.ok:
            return redirect(url_for('login.login_page'))

        user_data = user_info.json()
        
        email = user_data.get('email')
        name = user_data.get('name')
        profile_picture = user_data.get('picture')

        if not email or not name:
            return redirect(url_for('login.login_page'))

        # Check if user exists
        existing_user = find_user_by_email(email)

        if existing_user and existing_user.get('username') and existing_user.get('password'):
            # User exists and has completed signup - log them in
            
            # Update last login time
            update_user_login_time(existing_user['username'])
            
            # Set session data for existing user
            session['username'] = existing_user['username']
            session['email'] = existing_user['email']
            # Use default picture if no custom profile picture is set
            session['profile_picture'] = existing_user.get('profile_picture') or 'default_profile.png'
            session['user_id'] = existing_user['id']
            session['is_admin'] = False
            
            return redirect(url_for('login.landing_page'))
        else:
            # New user or Google user without username/password - redirect to signup2
            # Store Google info in session for signup2
            session['google_email'] = email
            session['google_name'] = name
            session['google_profile_picture'] = profile_picture
            
            return redirect(url_for('login.signup2_page'))

    except Exception as e:
        return redirect(url_for('login.login_page'))

# Password Reset Routes
@login_bp.route('/forgot-password', methods=['GET'])
def forgot_password_page():
    """Forgot password page"""
    # Clear any existing reset session to allow fresh requests
    session.pop('reset_email', None)
    session.pop('code_verified', None)
    return render_template('forget.html')

@login_bp.route('/forgot-password', methods=['POST'])
def forgot_password():
    """Handle forgot password form submission"""
    email = request.form.get('email')
    
    if not email:
        return redirect(url_for('login.forgot_password_page'))
    
    # Validate email format
    import re
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_regex, email):
        return redirect(url_for('login.forgot_password_page'))
    
    # Check if user exists
    user = find_user_by_email(email)
    print(f"DEBUG: User found for email {email}: {user is not None}")
    if user:
        print(f"DEBUG: User email_verified: {user.get('email_verified', False)}")
        print(f"DEBUG: User is_google_user: {user.get('is_google_user', False)}")
    
    if not user:
        flash('Email address not found. Please check your email or sign up.', 'error')
        return redirect(url_for('login.forgot_password_page'))
    
    # Check if email is verified
    if not user.get('email_verified', False):
        flash('Email not verified. Please verify your email first.', 'error')
        return redirect(url_for('login.forgot_password_page'))
    
    # Allow Google users to reset password if they have set one during signup2
    # (Google users who completed signup2 have passwords and should be able to reset them)
    
    # Generate and store reset code
    reset_code = generate_reset_code()
    print(f"DEBUG: Generated reset code: {reset_code} for email: {email}")
    
    if not store_reset_code(email, reset_code):
        print(f"DEBUG: Failed to store reset code for {email}")
        flash('Error generating reset code. Please try again.', 'error')
        return redirect(url_for('login.forgot_password_page'))
    
    print(f"DEBUG: Reset code stored successfully for {email}")
    
    # Send reset email
    if send_reset_email(email, reset_code):
        print(f"DEBUG: Reset email sent successfully to {email}")
        # Store email in session for verification step
        session['reset_email'] = email
        return redirect(url_for('login.verify_reset_code_page'))
    else:
        print(f"DEBUG: Failed to send reset email to {email}")
        flash('Error sending reset email. Please try again.', 'error')
        return redirect(url_for('login.forgot_password_page'))

@login_bp.route('/verify-code', methods=['GET'])
def verify_reset_code_page():
    """Verify reset code page"""
    if 'reset_email' not in session:
        return redirect(url_for('login.forgot_password_page'))
    
    return render_template('verify_code.html', email=session['reset_email'])

@login_bp.route('/verify-code', methods=['POST'])
def verify_reset_code():
    """Handle reset code verification"""
    if 'reset_email' not in session:
        return redirect(url_for('login.forgot_password_page'))
    
    email = session['reset_email']
    code = request.form.get('code', '').strip()
    
    if not code:
        flash('Please enter the verification code', 'error')
        return redirect(url_for('login.verify_reset_code_page'))
    
    # Verify the code against database
    is_valid, message = verify_reset_code_helper(email, code)
    
    if is_valid:
        # Ensure the user exists and can reset password (including Google users)
        user = find_user_by_email(email)
        if user:  # Allow all users (regular and Google) to proceed
            session['code_verified'] = True
            session.permanent = True
            return redirect(url_for('login.reset_password_page'))
        else:
            flash('User not found', 'error')
            return redirect(url_for('login.verify_reset_code_page'))
    else:
        flash(message or 'Invalid verification code', 'error')
        return redirect(url_for('login.verify_reset_code_page'))

@login_bp.route('/reset-password', methods=['GET'])
def reset_password_page():
    """Reset password page"""
    if 'reset_email' not in session or 'code_verified' not in session:
        return redirect(url_for('login.forgot_password_page'))
    
    return render_template('reset_password.html', email=session['reset_email'])

@login_bp.route('/reset-password', methods=['POST'])
def reset_password():
    """Handle password reset"""
    if 'reset_email' not in session or 'code_verified' not in session:
        return redirect(url_for('login.forgot_password_page'))
    
    email = session['reset_email']
    new_password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    if not new_password or not confirm_password:
        return redirect(url_for('login.reset_password_page'))
    
    if new_password != confirm_password:
        return redirect(url_for('login.reset_password_page'))
    
    # Validate password strength
    if len(new_password) < 8:
        return redirect(url_for('login.reset_password_page'))
    
    # Find user and update password (works for both regular and Google users)
    user = find_user_by_email(email)
    if not user:
        return redirect(url_for('login.forgot_password_page'))
    
    # Update password (Google users can now have/update passwords)
    new_password_hash = generate_password_hash(new_password)
    if update_user(user['id'], {'password': new_password_hash}):
        # Clear session data
        session.pop('reset_email', None)
        session.pop('code_verified', None)
        
        return redirect(url_for('login.login_page'))
    else:
        return redirect(url_for('login.reset_password_page'))

@login_bp.route('/resend-reset-code', methods=['POST'])
def resend_reset_code():
    """Resend reset code"""
    if 'reset_email' not in session:
        return jsonify({'success': False, 'message': 'Session expired. Please start over.'}), 400
    
    email = session['reset_email']
    
    # Generate new code
    reset_code = generate_reset_code()
    print(f"DEBUG RESEND: Generated new code {reset_code} for {email}")
    
    if not store_reset_code(email, reset_code):
        print(f"DEBUG RESEND: Failed to store new code for {email}")
        return jsonify({'success': False, 'message': 'Error generating new code. Please try again.'}), 500
    
    print(f"DEBUG RESEND: New code stored successfully for {email}")
    
    # Send reset email
    if send_reset_email(email, reset_code):
        print(f"DEBUG RESEND: New reset email sent successfully to {email}")
        return jsonify({'success': True, 'message': 'New verification code sent successfully!'})
    else:
        print(f"DEBUG RESEND: Failed to send new reset email to {email}")
        return jsonify({'success': False, 'message': 'Error sending email. Please try again.'}), 500

# Email Verification Routes
@login_bp.route('/verify-email', methods=['GET'])
def verify_email_page():
    """Email verification page"""
    email = request.args.get('email')
    
    if not email:
        return redirect(url_for('login.login_page'))
    
    # Check if verification code exists for this email
    connection = get_db_connection()
    if not connection:
        return redirect(url_for('login.login_page'))
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM verification_codes WHERE email = %s", (email,))
            code_data = cursor.fetchone()
            
        if not code_data:
            # Code might have been used already - redirect to login
            return redirect(url_for('login.login_page'))
    finally:
        connection.close()
    
    return render_template('verify-email.html', email=email)

@login_bp.route('/verify-email', methods=['POST'])
def verify_email():
    """Handle email verification code submission"""
    email = request.form.get('email')
    verification_code = request.form.get('verification_code')
    
    if not email or not verification_code:
        return redirect(url_for('login.verify_email_page', email=email))
    
    # Verify the code
    success, message, user_data = verify_email_code_helper(email, verification_code)
    print(f"DEBUG: Verification result - Success: {success}, Message: {message}")
    print(f"DEBUG: User data received: {user_data}")
    
    # Create user and redirect to login for non-Google users
    if user_data and not user_data.get('is_google_user', False):
        # This is a regular signup user - create account and go to login
        new_user = create_user(
            user_data['username'], 
            user_data['email'], 
            user_data['password'], 
            False,  # Not Google user
            user_data.get('profile_picture'),
            email_verified=True
        )
        # Redirect to login page
        return redirect('/login')
    
    # For other cases, redirect back to verification
    return redirect(url_for('login.verify_email_page', email=email))

@login_bp.route('/resend-code')
def resend_code():
    """Resend verification code"""
    email = request.args.get('email')
    if not email:
        return redirect(url_for('login.signup_page'))
    
    # Check if verification code exists for this email
    connection = get_db_connection()
    if not connection:
        return redirect(url_for('login.signup_page'))
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM verification_codes WHERE email = %s ORDER BY created_at DESC LIMIT 1", (email,))
            code_data = cursor.fetchone()
            
        if not code_data:
            return redirect(url_for('login.signup_page'))
        
        # Get existing user data
        import json
        user_data = json.loads(code_data['user_data']) if code_data['user_data'] else None
        if not user_data:
            return redirect(url_for('login.signup_page'))
    finally:
        connection.close()
    
    # Generate new verification code
    new_code = generate_verification_code()
    print(f"DEBUG RESEND EMAIL: Generated new verification code {new_code} for {email}")
    
    # Update the stored code
    if store_verification_code(email, new_code, user_data):
        print(f"DEBUG RESEND EMAIL: New verification code stored for {email}")
        # Send new verification email
        if send_verification_email(email, new_code):
            print(f"DEBUG RESEND EMAIL: New verification email sent to {email}")
        else:
            print(f"DEBUG RESEND EMAIL: Failed to send new verification email to {email}")
    else:
        print(f"DEBUG RESEND EMAIL: Failed to store new verification code for {email}")
    
    return redirect(url_for('login.verify_email_page', email=email))

# Debug Routes (remove in production)
@login_bp.route('/debug/codes/<email>')
def debug_codes(email):
    """Debug route to check codes in database"""
    connection = get_db_connection()
    if not connection:
        return f"Database connection failed"
    
    try:
        with connection.cursor() as cursor:
            # Check reset codes
            cursor.execute("SELECT * FROM reset_codes WHERE email = %s ORDER BY created_at DESC", (email,))
            reset_codes = cursor.fetchall()
            
            # Check verification codes  
            cursor.execute("SELECT * FROM verification_codes WHERE email = %s ORDER BY created_at DESC", (email,))
            verification_codes = cursor.fetchall()
            
            result = f"<h2>Debug: Codes for {email}</h2>"
            result += f"<h3>Reset Codes ({len(reset_codes)} found):</h3>"
            for code in reset_codes:
                result += f"<p>ID: {code['id']}, Code: {code['code']}, Expires: {code['expires']}, Attempts: {code['attempts']}</p>"
            
            result += f"<h3>Verification Codes ({len(verification_codes)} found):</h3>"
            for code in verification_codes:
                result += f"<p>ID: {code['id']}, Code: {code['code']}, Expires: {code['expires']}, Attempts: {code['attempts']}</p>"
            
            return result
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        connection.close()

# API Routes
@login_bp.route('/api/check-session', methods=['GET'])
def check_session():
    """Check if admin session is still valid"""
    is_admin = session.get('is_admin', False)
    admin_username = session.get('username')
    admin_logged_out = session.get('admin_logged_out', False)
    
    # If admin was logged out and trying to access via back button
    if admin_logged_out and session.get('admin_username') == ADMIN_USERNAME:
        # Restore admin session but require 2FA only if not already verified recently
        session['is_admin'] = True
        session['username'] = session['admin_username']
        
        # Check if 2FA was recently verified (within last 5 minutes)
        recent_2fa_verification = session.get('admin_2fa_verified_at')
        current_time = datetime.now(UTC).timestamp()
        
        if recent_2fa_verification and (current_time - recent_2fa_verification) < 300:  # 5 minutes
            # 2FA recently verified, allow access without requiring another 2FA
            session.pop('admin_logged_out', None)
            session.pop('admin_username', None)
            print(f"DEBUG SESSION CHECK: Admin back button detected, but 2FA recently verified, allowing access")
            return jsonify({'valid': True})
        else:
            # Require fresh 2FA
            session['admin_2fa_pending'] = True
            session.pop('admin_logged_out', None)
            session.pop('admin_username', None)
            print(f"DEBUG SESSION CHECK: Admin back button detected, requiring fresh 2FA")
            return jsonify({'valid': False, 'redirect': '/admin/2fa'})
    
    if is_admin and admin_username == ADMIN_USERNAME:
        # Admin session exists but check if 2FA is needed
        if session.get('admin_2fa_pending'):
            return jsonify({'valid': False, 'redirect': '/admin/2fa'})
        else:
            return jsonify({'valid': True})
    else:
        # Not an admin or no session
        return jsonify({'valid': False, 'redirect': '/login'})

@login_bp.route('/api/check-email', methods=['POST'])
def check_email():
    """Check if email is already taken"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip()
        
        if not email:
            return jsonify({'taken': False})
            
        user = find_user_by_email(email)
        return jsonify({'taken': user is not None})
    except Exception as e:
        return jsonify({'taken': False})

@login_bp.route('/api/security-tokens', methods=['GET'])
def get_security_tokens():
    """Generate security tokens for forms"""
    try:
        # Generate CSRF token
        csrf_token = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
        
        # Generate form token
        form_token = ''.join(random.choices(string.ascii_letters + string.digits, k=24))
        
        # Generate session ID if not exists
        if 'session_id' not in session:
            session['session_id'] = str(uuid.uuid4())
        
        # Store tokens in session for validation
        session['csrf_token'] = csrf_token
        session['form_token'] = form_token
        
        return jsonify({
            'success': True,
            'csrf_token': csrf_token,
            'form_token': form_token,
            'session_id': session['session_id'],
            'timestamp': str(int(datetime.now().timestamp()))
        })
    except Exception as e:
        print(f"Error generating security tokens: {e}")
        return jsonify({
            'success': False,
            'error': 'Failed to generate security tokens'
        }), 500

@login_bp.route('/api/captcha', methods=['GET'])
def get_captcha():
    """Generate CAPTCHA for signup form"""
    try:
        # Generate random CAPTCHA code
        chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
        captcha = ''.join(random.choices(chars, k=6))
        
        return jsonify({
            'success': True,
            'captcha': captcha
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': 'Failed to generate CAPTCHA'
        }), 500

@login_bp.route('/api/check-username', methods=['POST'])
def check_username_availability():
    """Check if username is available"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({
                'success': False,
                'available': False,
                'message': 'Username is required'
            })
        
        # Basic format validation
        if len(username) < 3:
            return jsonify({
                'success': True,
                'available': False,
                'message': 'Username too short (min 3 chars)'
            })
        
        if len(username) > 30:
            return jsonify({
                'success': True,
                'available': False,
                'message': 'Username too long (max 30 chars)'
            })
        
        if not re.match(r'^[a-zA-Z0-9_]+$', username):
            return jsonify({
                'success': True,
                'available': False,
                'message': 'Only letters, numbers, and underscores allowed'
            })
        
        if re.match(r'^[0-9]+$', username):
            return jsonify({
                'success': True,
                'available': False,
                'message': 'Username cannot be all numbers'
            })
        
        # Check if username already exists in database
        existing_user = find_user_by_username(username)
        if existing_user:
            return jsonify({
                'success': True,
                'available': False,
                'message': 'Username is already taken'
            })
        
        # Username is available
        return jsonify({
            'success': True,
            'available': True,
            'message': 'Username is available'
        })
        
    except Exception as e:
        print(f"Error checking username: {e}")
        return jsonify({
            'success': False,
            'available': False,
            'message': 'Error checking username'
        }), 500

# Admin Routes
@login_bp.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard"""
    # Get statistics
    stats = get_user_stats()
    
    # Get recent users (last 10)
    users = load_users()
    recent_users = sorted(users, key=lambda x: x.get('created_at', ''), reverse=True)[:10]
    
    # Get all users for the table
    all_users = sorted(users, key=lambda x: x.get('created_at', ''), reverse=True)
    
    return render_template('admin_dashboard.html', 
                         total_users=stats['total_users'],
                         google_users=stats['google_users'],
                         regular_users=stats['regular_users'],
                         users_today=stats['users_today'],
                         users_this_week=stats['users_this_week'],
                         recent_users=recent_users,
                         all_users=all_users)

@login_bp.route('/admin/user/<user_id>')
@admin_required
def admin_user_detail(user_id):
    """Admin user detail page"""
    user = find_user_by_id(user_id)
    if not user:
        return redirect(url_for('login.admin_dashboard'))
    return render_template('admin_user_detail.html', user=user)

@login_bp.route('/admin/user/<user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    """Admin delete user"""
    user = find_user_by_id(user_id)
    if user:
        delete_user_by_id(user_id)
    return redirect(url_for('login.admin_dashboard'))

@login_bp.route('/admin/users/export')
@admin_required
def admin_export_users():
    """Export users to CSV"""
    import csv
    from io import StringIO
    from flask import make_response
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['ID', 'Username', 'Email', 'Is Google User', 'Occupation', 
                     'Birthday', 'Labels', 'Online Start', 'Online End', 
                     'Created At', 'Last Login'])
    
    # Write user data
    users = load_users()
    for user in users:
        writer.writerow([
            user.get('id', ''),
            user.get('username', ''),
            user.get('email', ''),
            user.get('is_google_user', False),
            user.get('occupation', ''),
            user.get('birthday', ''),
            user.get('labels', ''),
            user.get('online_start_time', ''),
            user.get('online_end_time', ''),
            user.get('created_at', ''),
            user.get('last_login', '')
        ])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.csv'
    
    return response

@login_bp.route('/admin/users/export/excel')
@admin_required
def admin_export_users_excel():
    """Export users to Excel format"""
    from flask import make_response
    import io
    
    try:
        # Try to import openpyxl for Excel functionality
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users Export"
        
        # Define headers
        headers = ['ID', 'Username', 'Email', 'Is Google User', 'Email Verified', 'Occupation', 
                  'Birthday', 'Labels', 'Online Start', 'Online End', 
                  'Created At', 'Last Login']
        
        # Write headers with formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write user data
        users = load_users()
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.get('id', ''))
            ws.cell(row=row, column=2, value=user.get('username', ''))
            ws.cell(row=row, column=3, value=user.get('email', ''))
            ws.cell(row=row, column=4, value='Yes' if user.get('is_google_user', False) else 'No')
            ws.cell(row=row, column=5, value='Yes' if user.get('email_verified', False) else 'No')
            ws.cell(row=row, column=6, value=user.get('occupation', ''))
            ws.cell(row=row, column=7, value=user.get('birthday', ''))
            ws.cell(row=row, column=8, value=user.get('labels', ''))
            ws.cell(row=row, column=9, value=user.get('online_start_time', ''))
            ws.cell(row=row, column=10, value=user.get('online_end_time', ''))
            ws.cell(row=row, column=11, value=str(user.get('created_at', '')))
            ws.cell(row=row, column=12, value=str(user.get('last_login', '')))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
        
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        flash('Excel export requires openpyxl package. Exporting as CSV instead.', 'warning')
        return redirect(url_for('login.admin_export_users'))

# Admin 2FA Routes
@login_bp.route('/admin/2fa')
def admin_2fa_page():
    """Admin 2FA verification page"""
    if not session.get('admin_2fa_pending'):
        return redirect(url_for('login.login_page'))
    
    # Check if this is a re-verification after logout (admin is already in session)
    is_reverification = session.get('is_admin') and session.get('username') == ADMIN_USERNAME
    
    return render_template('admin_2fa.html', is_reverification=is_reverification)

@login_bp.route('/admin/2fa/verify', methods=['POST'])
def admin_2fa_verify():
    """Verify 2FA code and complete admin login"""
    if not session.get('admin_2fa_pending'):
        return jsonify({'success': False, 'message': 'No 2FA session found'}), 400
    
    try:
        data = request.get_json()
        code = data.get('code', '').strip()
        
        valid_codes = [
            '090808@DSF',
            '071008@DSF', 
            '170304@DSF',
            '240301@DSF',
            '240302@DSF',
            '240303@DSF'
        ]
        
        if code in valid_codes:
            # 2FA successful - clear pending flag and set short-lived access token
            session.pop('admin_2fa_pending', None)
            session['admin_2fa_verified_at'] = datetime.now(UTC).timestamp()
            return jsonify({'success': True, 'message': '2FA verification successful'})
        else:
            return jsonify({'success': False, 'message': 'Invalid 2FA code'}), 400
            
    except Exception as e:
        print(f"Error verifying 2FA: {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

@login_bp.route('/admin/security-alert', methods=['POST'])
def admin_security_alert():
    """Handle 2FA security breach notifications"""
    try:
        data = request.get_json()
        
        if data and data.get('event') == '2FA_BREACH_ATTEMPT':
            # Send security alert email
            success = send_security_breach_email(
                ADMIN_EMAIL, 
                data.get('attempts', 0),
                data.get('timestamp', '')
            )
            
            if success:
                return jsonify({'success': True, 'message': 'Security alert sent'})
            else:
                return jsonify({'success': False, 'message': 'Failed to send alert'}), 500
        
        return jsonify({'success': False, 'message': 'Invalid request'}), 400
        
    except Exception as e:
        print(f"Error handling security alert: {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

def send_security_breach_email(admin_email, attempts, timestamp):
    """Send security breach notification email"""
    try:
        # Get Gmail credentials from environment variables
        gmail_user = os.getenv('GMAIL_USER')
        gmail_password = os.getenv('GMAIL_APP_PASSWORD')
        
        if not gmail_user or not gmail_password:
            print("Gmail credentials not found for security alert")
            return False
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"CultureQuest Security <{gmail_user}>"
        msg['To'] = admin_email
        msg['Subject'] = "🚨 SECURITY ALERT: Admin 2FA Breach Attempt - CultureQuest"
        
        # Email body with security alert
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px; background: #dc3545; color: white; padding: 20px; border-radius: 8px;">
                    <h1 style="margin: 0; font-size: 24px;">🚨 SECURITY ALERT</h1>
                    <p style="margin: 10px 0 0; font-size: 16px;">Possible Security Breach Detected</p>
                </div>
                
                <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #dc3545;">
                    <h2 style="color: #dc3545; margin-top: 0;">Admin 2FA Access Attempt Blocked</h2>
                    <p><strong>Event:</strong> Multiple failed 2FA attempts on admin account</p>
                    <p><strong>Failed Attempts:</strong> {attempts} consecutive failures</p>
                    <p><strong>Timestamp:</strong> {timestamp}</p>
                    <p><strong>Status:</strong> ❌ Access DENIED - Account temporarily locked</p>
                </div>
                
                <div style="background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="color: #856404; margin-top: 0;">⚠️ Immediate Actions Required</h3>
                    <ul style="color: #856404; margin: 10px 0;">
                        <li>Review server access logs immediately</li>
                        <li>Check for suspicious network activity</li>
                        <li>Verify admin account security</li>
                        <li>Consider changing admin credentials if compromise is suspected</li>
                        <li>Monitor system for unusual activity</li>
                    </ul>
                </div>
                
                <div style="background: #d1ecf1; border: 1px solid #bee5eb; padding: 15px; border-radius: 8px; margin-bottom: 20px;">
                    <h3 style="color: #0c5460; margin-top: 0;">🔒 Security Details</h3>
                    <p style="color: #0c5460; margin: 5px 0;"><strong>System:</strong> CultureQuest Admin Panel</p>
                    <p style="color: #0c5460; margin: 5px 0;"><strong>Protection:</strong> 2FA Authentication System</p>
                    <p style="color: #0c5460; margin: 5px 0;"><strong>Action:</strong> Automated lockout after 5 failed attempts</p>
                </div>
                
                <div style="text-align: center; padding-top: 20px; border-top: 1px solid #eee; color: #666; font-size: 12px;">
                    <p><strong>This is an automated security notification from CultureQuest</strong></p>
                    <p>Please do not reply to this email. Contact your system administrator if you need assistance.</p>
                    <p>Generated at: {timestamp}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.attach(MIMEText(body, 'html'))
        
        # Gmail SMTP setup
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, gmail_password)
        
        # Send email
        text = msg.as_string()
        server.sendmail(gmail_user, admin_email, text)
        server.quit()
        
        print(f"Security breach alert sent to {admin_email}")
        return True
        
    except Exception as e:
        print(f"Failed to send security breach email: {e}")
        return False

# Initialize database when blueprint is imported
try:
    init_database()
except Exception as e:
    print(f"Warning: Could not initialize database on import: {e}")

# No error handlers - let Flask handle errors naturally